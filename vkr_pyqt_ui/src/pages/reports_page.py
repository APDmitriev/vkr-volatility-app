from __future__ import annotations

from datetime import datetime
from html import escape
from pathlib import Path
from typing import Any
from zipfile import ZIP_DEFLATED, ZipFile

from PyQt6.QtGui import QTextDocument
from PyQt6.QtPrintSupport import QPrinter
from PyQt6.QtWidgets import (
    QCheckBox,
    QFileDialog,
    QHBoxLayout,
    QMessageBox,
    QPushButton,
    QTextEdit,
    QVBoxLayout,
)

from src.services.app_state import AppState
from src.services.backend_service import BackendService
from src.widgets.common import BasePage, Card


class ReportsPage(BasePage):
    """Страница формирования отчётов.

    Предпросмотр строится только из выбранных чекбоксами разделов.
    Экспорт поддерживает PDF, DOCX и XLSX без обращения к backend.
    """

    def __init__(self, state: AppState, backend: BackendService) -> None:
        super().__init__("Отчёты")
        self.state = state
        self.backend = backend
        self.checkboxes: dict[str, QCheckBox] = {}

        options_card = Card("Выбор содержимого")
        checks_layout = QVBoxLayout()
        for key, title in [
            ("data", "Данные"),
            ("analysis", "Анализ"),
            ("models", "Модели"),
            ("forecast", "Прогноз"),
        ]:
            checkbox = QCheckBox(title)
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(self.update_preview)
            self.checkboxes[key] = checkbox
            checks_layout.addWidget(checkbox)
        options_card.layout.addLayout(checks_layout)
        self.add_card(options_card)

        preview_card = Card("Предпросмотр")
        self.report_preview = QTextEdit()
        self.report_preview.setReadOnly(True)
        self.report_preview.setMinimumHeight(330)
        preview_card.layout.addWidget(self.report_preview)
        self.add_card(preview_card)

        export_card = Card("Экспорт")
        export_row = QHBoxLayout()
        self.pdf_btn = QPushButton("PDF")
        self.docx_btn = QPushButton("DOCX")
        self.xlsx_btn = QPushButton("XLSX")
        export_row.addWidget(self.pdf_btn)
        export_row.addWidget(self.docx_btn)
        export_row.addWidget(self.xlsx_btn)
        export_row.addStretch()
        export_card.layout.addLayout(export_row)
        self.add_card(export_card)
        self.root_layout.addStretch()

        self.pdf_btn.clicked.connect(self.export_pdf)
        self.docx_btn.clicked.connect(self.export_docx)
        self.xlsx_btn.clicked.connect(self.export_xlsx)
        self.state.dataset_changed.connect(lambda _: self.update_preview())
        self.state.project_changed.connect(lambda _: self.update_preview())
        self.state.experiments_changed.connect(self.update_preview)
        self.update_preview()

    def selected_sections(self) -> list[str]:
        return [key for key, checkbox in self.checkboxes.items() if checkbox.isChecked()]

    def update_preview(self) -> None:
        self.report_preview.setPlainText(self.build_plain_report())

    def build_plain_report(self) -> str:
        sections = self.selected_sections()
        if not sections:
            return "Выберите хотя бы один раздел отчёта."

        lines: list[str] = [
            "Отчёт по анализу и прогнозированию временного ряда",
            f"Дата формирования: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            "",
        ]

        if "data" in sections:
            lines.extend(self._data_section())
        if "analysis" in sections:
            lines.extend(self._analysis_section())
        if "models" in sections:
            lines.extend(self._models_section())
        if "forecast" in sections:
            lines.extend(self._forecast_section())

        return "\n".join(lines).rstrip()

    def build_report_rows(self) -> list[tuple[str, str, str]]:
        """Табличное представление для XLSX."""
        rows: list[tuple[str, str, str]] = []
        for section in self.build_plain_report().split("\n\n"):
            section = section.strip()
            if not section:
                continue
            lines = section.splitlines()
            title = lines[0]
            rows.append((title, "", ""))
            for line in lines[1:]:
                if ":" in line:
                    key, value = line.split(":", 1)
                    rows.append(("", key.strip(), value.strip()))
                else:
                    rows.append(("", line.strip(), ""))
            rows.append(("", "", ""))
        return rows

    def _data_section(self) -> list[str]:
        df = self.state.current_df
        processed = self.state.processed_df
        columns = list(df.columns) if df is not None else []
        processed_columns = list(processed.columns) if processed is not None else []
        return [
            "1. Данные",
            f"Проект: {self.state.current_project}",
            f"Датасет: {self.state.current_dataset_name}",
            f"Количество строк: {len(df) if df is not None else '—'}",
            f"Количество столбцов: {len(columns) if columns else '—'}",
            f"Подготовленный датасет: {'да' if processed is not None else 'нет'}",
            f"Столбцы после предобработки: {', '.join(map(str, processed_columns[:8])) if processed_columns else '—'}",
            "",
        ]

    def _analysis_section(self) -> list[str]:
        analysis = self.state.last_analysis or {}
        profile = self.state.last_volatility_profile or analysis.get("volatility_profile", {}) or {}
        statistics = analysis.get("statistics", {}) or analysis.get("summary", {}) or {}
        lines = [
            "2. Анализ",
            f"Рекомендуемая модель: {profile.get('recommended_model', self.state.recommended_model or '—')}",
            f"Уровень волатильности: {profile.get('volatility_level', '—')}",
            f"Сезонность: {self._yes_no(profile.get('seasonality_detected'))}",
            f"Кластеризация волатильности: {self._yes_no(profile.get('volatility_clustering_detected'))}",
            f"Автокорреляция: {self._yes_no(profile.get('autocorrelation_detected'))}",
        ]
        if statistics:
            for key in ["count", "mean", "std", "min", "max", "median"]:
                if key in statistics:
                    lines.append(f"{key}: {self._fmt(statistics.get(key))}")
        lines.append("")
        return lines

    def _models_section(self) -> list[str]:
        experiments = list(self.state.experiments or [])
        lines = [
            "3. Модели и эксперименты",
            f"Количество сохранённых экспериментов: {len(experiments)}",
        ]
        if not experiments:
            lines.append("Эксперименты отсутствуют.")
            lines.append("")
            return lines

        for index, exp in enumerate(experiments[:8], start=1):
            model = exp.get("model", "—")
            dataset = exp.get("dataset_name", "—")
            rmse = self._fmt(exp.get("rmse"))
            mae = self._fmt(exp.get("mae"))
            mape = self._fmt(exp.get("mape"))
            lines.append(f"{index}. {model} | датасет: {dataset} | RMSE: {rmse} | MAE: {mae} | MAPE: {mape}")
        lines.append("")
        return lines

    def _forecast_section(self) -> list[str]:
        forecast = self.state.last_forecast or {}
        model = forecast.get("model", "—")
        next_value = forecast.get("next_value", "—")
        horizon = len(forecast.get("forecast", []) or []) if forecast else "—"
        metrics = forecast.get("metrics", {}) if isinstance(forecast.get("metrics", {}), dict) else {}
        lines = [
            "4. Прогноз",
            f"Модель: {model}",
            f"Горизонт прогноза: {horizon}",
            f"Следующее прогнозное значение: {self._fmt(next_value)}",
        ]
        if metrics:
            for key in ["mae", "mse", "rmse", "mape"]:
                if key in metrics:
                    lines.append(f"{key.upper()}: {self._fmt(metrics.get(key))}")
        lines.append("")
        return lines

    def export_pdf(self) -> None:
        file_path = self._choose_path("PDF (*.pdf)", ".pdf")
        if not file_path:
            return
        try:
            document = QTextDocument()
            document.setHtml(self._report_html())
            printer = QPrinter(QPrinter.PrinterMode.HighResolution)
            printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
            printer.setOutputFileName(str(file_path))
            document.print(printer)
            self._export_done(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Ошибка экспорта", str(exc))

    def export_docx(self) -> None:
        file_path = self._choose_path("Word (*.docx)", ".docx")
        if not file_path:
            return
        try:
            self._write_docx(file_path, self.build_plain_report())
            self._export_done(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Ошибка экспорта", str(exc))

    def export_xlsx(self) -> None:
        file_path = self._choose_path("Excel (*.xlsx)", ".xlsx")
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Отчёт"
            ws.append(["Раздел", "Показатель", "Значение"])
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="D9EAF7")

            for row in self.build_report_rows():
                ws.append(list(row))
                if row[0] and not row[1] and not row[2]:
                    for cell in ws[ws.max_row]:
                        cell.font = Font(bold=True)

            for column_index, width in enumerate([36, 42, 70], start=1):
                ws.column_dimensions[get_column_letter(column_index)].width = width
            wb.save(file_path)
            self._export_done(file_path)
        except Exception as exc:
            QMessageBox.critical(self, "Ошибка экспорта", str(exc))

    def _choose_path(self, file_filter: str, suffix: str) -> Path | None:
        default_name = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}{suffix}"
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить отчёт", default_name, file_filter)
        if not file_path:
            return None
        path = Path(file_path)
        if path.suffix.lower() != suffix:
            path = path.with_suffix(suffix)
        return path

    def _report_html(self) -> str:
        paragraphs = []
        for line in self.build_plain_report().splitlines():
            safe = escape(line)
            if not safe:
                paragraphs.append("<br>")
            elif safe[0].isdigit() and ". " in safe[:4]:
                paragraphs.append(f"<h2>{safe}</h2>")
            elif safe.startswith("Отчёт"):
                paragraphs.append(f"<h1>{safe}</h1>")
            else:
                paragraphs.append(f"<p>{safe}</p>")
        return """
        <html>
        <head>
            <meta charset=\"utf-8\">
            <style>
                body { font-family: Arial, sans-serif; font-size: 11pt; }
                h1 { color: #0b2d63; font-size: 18pt; }
                h2 { color: #0b2d63; font-size: 14pt; margin-top: 18px; }
                p { margin: 4px 0; }
            </style>
        </head>
        <body>
        """ + "\n".join(paragraphs) + "\n</body></html>"

    @staticmethod
    def _write_docx(path: Path, text: str) -> None:
        paragraphs = "".join(
            f"<w:p><w:r><w:t>{escape(line) if line else ''}</w:t></w:r></w:p>"
            for line in text.splitlines()
        )
        content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
<Default Extension="xml" ContentType="application/xml"/>
<Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
        rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
        document = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
<w:body>{paragraphs}<w:sectPr/></w:body>
</w:document>"""
        with ZipFile(path, "w", ZIP_DEFLATED) as docx:
            docx.writestr("[Content_Types].xml", content_types)
            docx.writestr("_rels/.rels", rels)
            docx.writestr("word/document.xml", document)

    def _export_done(self, path: Path) -> None:
        self.state.set_status(f"Отчёт экспортирован: {path}")

    @staticmethod
    def _fmt(value: Any) -> str:
        if value in (None, "", "—"):
            return "—"
        try:
            number = float(value)
        except (TypeError, ValueError):
            return str(value)
        if abs(number) >= 1000:
            return f"{number:.2f}"
        if abs(number) >= 1:
            return f"{number:.4f}"
        return f"{number:.6f}"

    @staticmethod
    def _yes_no(value: Any) -> str:
        if value is True:
            return "да"
        if value is False:
            return "нет"
        return "—"
